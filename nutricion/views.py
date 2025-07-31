import base64
from datetime import date, datetime
import io
import os
import unicodedata
import pyexcel as pe
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from venv import logger
from django.views.generic import DeleteView, ListView, View, TemplateView, UpdateView, DetailView
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.contrib.auth.decorators import permission_required
from django.urls import reverse, reverse_lazy
from django.shortcuts import get_object_or_404, redirect, render
from django.contrib import messages
from django.http import Http404, HttpResponse, HttpResponseForbidden, JsonResponse
from matplotlib.dates import relativedelta
from reversion.models import Version
from django.db.models import OuterRef, Subquery, Max
from reversion import create_revision, set_user, set_comment
from django.utils.decorators import method_decorator
from django.views.decorators.http import require_GET
from django.utils.dateformat import DateFormat
from django.core.paginator import Paginator
from django.shortcuts import get_object_or_404
from django.template.loader import render_to_string
from django.http import HttpResponse
from weasyprint import CSS, HTML
from weasyprint.text.fonts import FontConfiguration
import json
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from django.db.models import Exists, OuterRef


from sistema_nutricion import settings
from .utils import BaseDeleteViewConCancel, FiltroCAIMixin, obtener_desviacion_oms_peso_edad, obtener_desviacion_oms_peso_talla, obtener_desviacion_oms_talla_edad
from .models import CAI_CHOICES, Paciente, Trabajador, SeguimientoTrimestral, SeguimientoTrabajador, Usuario, OmsPesoEdad, OmsPesoTalla, OmsTallaEdad
from .forms import (
    PacienteForm, TrabajadorForm,
    SeguimientoTrimestralForm, SeguimientoTrabajadorForm,
    UsuarioCreacionForm, UsuarioEdicionForm
)
from .utils import RevisionCreateView, RevisionUpdateView, CancelUrlMixin, InitialFromModelMixin

#-------------------
# Views Administración 
#-------------------

class AdminRequiredMixin(UserPassesTestMixin):
    def test_func(self):
        return self.request.user.is_authenticated and self.request.user.is_admin

class HomeView(LoginRequiredMixin, View):
    CAI_CHOICES = [
        ("ESTEFANIA CASTANEDA NUNEZ", "CAI Estefanía Castañeda Núñez"),
        ("JOSEFINA VICENS", "CAI Josefina Vicens"),
        ("JULIETA CAMPOS DE GONZALEZ PEDRERO", "CAI Julieta Campos de González Pedrero"),
        ("JOSE MARIA PINO SUAREZ", "CAI José María Pino Suárez"),
        ("MARINA CORTAZAR VDA. DE ESCOBAR", "CAI Marina Cortázar Viuda de Escobar"),
        ("EVA SAMANO DE LOPEZ MATEOS", "CAI Eva Sámano de López Mateos"),
    ]

    def normalizar(self, texto):
        texto = ''.join(c for c in unicodedata.normalize('NFD', texto or '') if unicodedata.category(c) != 'Mn')
        return texto.replace('.', '').upper().strip()

    def filtrar_por_cai_normalizado(self, modelo, clave_norm):
        return modelo.objects.filter(
            is_deleted=False,
            cai__isnull=False
        ).filter(
            cai__in=[
                cai for cai in modelo.objects.values_list('cai', flat=True)
                if self.normalizar(cai) == clave_norm
            ]
        )

    def get(self, request):
        user_cai = request.user.cai if request.user.is_nutriologo else None
        cai_dict = {self.normalizar(k): v for k, v in self.CAI_CHOICES}

        if user_cai:
            user_cai_norm = self.normalizar(user_cai)
            clave_oficial = next((k for k in cai_dict if self.normalizar(k) == user_cai_norm), user_cai)
            clave_norm = self.normalizar(clave_oficial)

            total_ninos = self.filtrar_por_cai_normalizado(Paciente, clave_norm).count()
            total_trabajadores = self.filtrar_por_cai_normalizado(Trabajador, clave_norm).count()
            total_seguimientos = (
                SeguimientoTrimestral.objects.filter(
                    is_deleted=False,
                    paciente__is_deleted=False,
                    paciente__cai__in=[
                        cai for cai in Paciente.objects.values_list('cai', flat=True)
                        if self.normalizar(cai) == clave_norm
                    ]
                ).count() +
                SeguimientoTrabajador.objects.filter(
                    is_deleted=False,
                    trabajador__is_deleted=False,
                    trabajador__cai__in=[
                        cai for cai in Trabajador.objects.values_list('cai', flat=True)
                        if self.normalizar(cai) == clave_norm
                    ]
                ).count()
            )
            cais = [(clave_oficial, cai_dict.get(self.normalizar(clave_oficial), clave_oficial))]
        else:
            total_ninos = Paciente.objects.filter(is_deleted=False).count()
            total_trabajadores = Trabajador.objects.filter(is_deleted=False).count()
            total_seguimientos = (
                SeguimientoTrimestral.objects.filter(
                    is_deleted=False,
                    paciente__is_deleted=False
                ).count() +
                SeguimientoTrabajador.objects.filter(
                    is_deleted=False,
                    trabajador__is_deleted=False
                ).count()
            )
            cais = self.CAI_CHOICES

        detalle_cais = {}
        for clave, nombre in cais:
            clave_norm = self.normalizar(clave)

            trabajadores = self.filtrar_por_cai_normalizado(Trabajador, clave_norm).values('id', 'nombre', 'cargo')
            ninos = self.filtrar_por_cai_normalizado(Paciente, clave_norm).values('id', 'nombre')

            seguimientos_ninos = SeguimientoTrimestral.objects.filter(
                is_deleted=False,
                paciente__is_deleted=False,
                paciente__cai__in=[
                    cai for cai in Paciente.objects.values_list('cai', flat=True)
                    if self.normalizar(cai) == clave_norm
                ]
            ).count()

            seguimientos_trabajadores = SeguimientoTrabajador.objects.filter(
                is_deleted=False,
                trabajador__is_deleted=False,
                trabajador__cai__in=[
                    cai for cai in Trabajador.objects.values_list('cai', flat=True)
                    if self.normalizar(cai) == clave_norm
                ]
            ).count()

            detalle_cais[clave] = {
                'nombre': nombre,
                'trabajadores': list(trabajadores),
                'ninos': list(ninos),
                'total_trabajadores': len(trabajadores),
                'total_ninos': len(ninos),
                'total_seguimientos': seguimientos_ninos + seguimientos_trabajadores
            }

        return render(request, 'home.html', {
            'usuario': request.user,
            'total_ninos': total_ninos,
            'total_trabajadores': total_trabajadores,
            'total_seguimientos': total_seguimientos,
            'cais': cais,
            'detalle_cais': detalle_cais,
        })
    
#-------------------
# Views Usuarios
#-------------------

class GestionUsuariosView(LoginRequiredMixin, AdminRequiredMixin, View):
    def get(self, request):
        usuarios = Usuario.objects.all().order_by('username')
        editar = False
        user_id = request.GET.get('edit')
        usuario_a_editar = None

        # Aplicar filtros
        rol_filter = request.GET.get('rol')
        cai_filter = request.GET.get('cai')
        
        if rol_filter:
            usuarios = usuarios.filter(rol=rol_filter)
        if cai_filter:
            usuarios = usuarios.filter(cai=cai_filter)

        if user_id:
            usuario_a_editar = get_object_or_404(Usuario, id=user_id)
            form = UsuarioEdicionForm(instance=usuario_a_editar)
            editar = True
        else:
            form = UsuarioCreacionForm()

        return render(request, 'gestionar_usuarios.html', {
            'usuarios': usuarios,
            'form': form,
            'editar': editar,
            'usuario': usuario_a_editar,
            'cai_choices': CAI_CHOICES,
            'filtro_rol': rol_filter,
            'filtro_cai': cai_filter
        })

    def post(self, request):
        user_id = request.GET.get('edit')
        editar = bool(user_id)

        if editar:
            usuario = get_object_or_404(Usuario, id=user_id)
            form = UsuarioEdicionForm(request.POST, instance=usuario)
        else:
            form = UsuarioCreacionForm(request.POST)

        if form.is_valid():
            usuario = form.save(commit=False)
            
            # Manejo especial del CAI
            if form.cleaned_data.get('rol') == 'nutriologo':
                usuario.cai = form.cleaned_data.get('cai')
            else:
                usuario.cai = None
            
            usuario.save()
            
            mensaje = '[usuarios] Usuario actualizado exitosamente.' if editar else '[usuarios] Usuario agregado exitosamente.'
            messages.success(request, mensaje)
            return redirect('gestionar_usuarios')

        # Si hay errores, volver a mostrar el formulario
        usuarios = Usuario.objects.all().order_by('username')
        return render(request, 'gestionar_usuarios.html', {
            'usuarios': usuarios,
            'form': form,
            'editar': editar,
            'usuario': usuario if editar else None,
            'cai_choices': CAI_CHOICES,
            'filtro_rol': request.GET.get('rol'),
            'filtro_cai': request.GET.get('cai')
        })

class EliminarUsuarioView(LoginRequiredMixin, AdminRequiredMixin, DeleteView):
    model = Usuario
    template_name = 'eliminar_usuario.html'
    success_url = reverse_lazy('gestionar_usuarios')

    def form_valid(self, form):
        if self.object.is_superuser:
            messages.error(self.request, '[usuarios] No puedes eliminar a un superusuario.')
            return redirect('gestionar_usuarios')
        messages.success(self.request, '[usuarios] Usuario eliminado correctamente.')
        return super().form_valid(form)
    
#-------------------
# Views Niños
#-------------------

class RegistroNinoView(FiltroCAIMixin, RevisionCreateView):
    model = Paciente
    form_class = PacienteForm
    template_name = 'registro_ninos.html'
    success_url = reverse_lazy('registro_ninos')
    comment = "Registro de nuevo niño"
    success_message = "[ninos] Niño agregado exitosamente."

    def form_valid(self, form):
        # Asignar CAI automáticamente si es nutriólogo
        if self.request.user.is_nutriologo:
            form.instance.cai = self.request.user.cai

        form.instance.created_by = self.request.user  # Auditoría
        self.object = form.save()
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        ultimos_ninos = Paciente.objects.filter(is_deleted=False)
        if self.request.user.is_nutriologo and self.request.user.cai:
            ultimos_ninos = ultimos_ninos.filter(cai=self.request.user.cai)
        context['ultimos_ninos'] = ultimos_ninos.order_by('-id')[:5]

        # Info IMC para mostrar si se requiere
        context['info_imc'] = {
            'menores_5': {
                'bajo_peso': '<14.0',
                'normal': '14.0-16.9',
                'sobrepeso': '17.0-17.9',
                'obesidad': '≥18.0'
            },
            'mayores_5': {
                'bajo_peso': 'Percentil <5',
                'normal': 'Percentil 5-85',
                'sobrepeso': 'Percentil 85-95',
                'obesidad': 'Percentil ≥95'
            }
        }

        return context

class EditarNinoView(CancelUrlMixin, RevisionUpdateView):
    model = Paciente
    form_class = PacienteForm
    template_name = 'editar_nino.html'
    comment = "Edición de niño"
    success_message = "[ninos] Niño actualizado correctamente."
    cancel_url_default = 'registro_ninos'
    cancel_url_alt = 'historial'

    def form_valid(self, form):
        form.instance.updated_by = self.request.user  # Auditoría
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy(self.cancel_url_alt if 'from' in self.request.GET else self.cancel_url_default)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['cancel_url'] = self.request.GET.get('from', 'lista_pacientes')
        return context

    def get_object(self):
        return get_object_or_404(Paciente, pk=self.kwargs['pk'], is_deleted=False)

class EliminarNinoView(LoginRequiredMixin, View):
    model = Paciente
    template_name = 'eliminar_nino.html'

    def dispatch(self, request, *args, **kwargs):
        paciente = get_object_or_404(Paciente, pk=self.kwargs['pk'])

        # Permitir a admins, jefas de departamento o nutriólogos del mismo CAI
        if not (
            request.user.is_admin
            or request.user.is_jefe_departamento
            or (request.user.is_nutriologo and paciente.cai == request.user.cai)
        ):
            messages.error(request, '[ninos] No tienes permiso para eliminar este niño.')
            return redirect('registro_ninos' if 'from' not in request.GET else 'historial')

        return super().dispatch(request, *args, **kwargs)

    def get_object(self):
        return get_object_or_404(Paciente, pk=self.kwargs['pk'])

    def get(self, request, *args, **kwargs):
        self.object = self.get_object()
        context = self.get_context_data()
        return render(request, self.template_name, context)

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        self.object.deleted_by = request.user  # Auditoría
        self.object.is_deleted = True
        self.object.save()
        messages.success(self.request, '[ninos] Niño eliminado correctamente.')
        return redirect(self.get_success_url())

    def get_success_url(self):
        return reverse_lazy('registro_ninos' if 'from' not in self.request.GET else 'historial')

    def get_context_data(self, **kwargs):
        return {
            'cancel_url': 'historial' if 'from' in self.request.GET else 'registro_ninos',
            'nino': self.get_object()
        }
    
class RestaurarNinoView(LoginRequiredMixin, View):
    model = Paciente
    template_name = 'restaurar_nino.html'

    def dispatch(self, request, *args, **kwargs):
        paciente = get_object_or_404(Paciente, pk=self.kwargs['pk'], is_deleted=True)

        # Solo admins o jefas de departamento pueden restaurar
        if not (request.user.is_admin or request.user.is_jefe_departamento):
            messages.error(request, '[ninos] No tienes permiso para restaurar este niño.')
            return redirect('registro_ninos' if 'from' not in request.GET else 'historial')

        return super().dispatch(request, *args, **kwargs)

    def get(self, request, *args, **kwargs):
        paciente = self.get_object()
        return render(request, self.template_name, {'nino': paciente})

    def post(self, request, *args, **kwargs):
        paciente = self.get_object()
        paciente.is_deleted = False
        paciente.save()
        messages.success(request, 'Niño restaurado correctamente.')
        return redirect(self.get_success_url())

    def get_object(self):
        return get_object_or_404(Paciente, pk=self.kwargs['pk'], is_deleted=True)

    def get_success_url(self):
        return reverse_lazy('registro_ninos')
    
class ListaPacientesEliminadosView(LoginRequiredMixin, ListView):
    template_name = 'pacientes_eliminados.html'
    model = Paciente
    context_object_name = 'pacientes'

    def get_queryset(self):
        # Solo mostrar eliminados
        qs = Paciente.objects.filter(is_deleted=True)
        if self.request.user.is_nutriologo:
            return qs.filter(cai=self.request.user.cai)
        return qs
    
#-------------------
# Views Trabajadores
#-------------------

class RegistroTrabajadorView(FiltroCAIMixin, RevisionCreateView):
    model = Trabajador
    form_class = TrabajadorForm
    template_name = 'registro_trabajadores.html'
    success_url = reverse_lazy('registro_trabajadores')
    success_message = '[trabajadores] Trabajador registrado exitosamente.'
    comment = "Registro de nuevo trabajador"

    def form_valid(self, form):
        if self.request.user.is_nutriologo:
            form.instance.cai = self.request.user.cai
        form.instance.created_by = self.request.user # Auditoría
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        ultimos_trabajadores = Trabajador.objects.filter(is_deleted=False)
        if self.request.user.is_nutriologo and self.request.user.cai:
            ultimos_trabajadores = ultimos_trabajadores.filter(cai=self.request.user.cai)
        context['ultimos_trabajadores'] = ultimos_trabajadores.order_by('-id')[:5]
        return context

class EditarTrabajadorView(CancelUrlMixin, RevisionUpdateView):
    model = Trabajador
    form_class = TrabajadorForm
    template_name = 'editar_trabajador.html'
    success_message = "Trabajador actualizado correctamente"
    comment = "Edición de trabajador"
    cancel_url_default = 'historial'
    cancel_url_alt = 'historial'

    def form_valid(self, form):
        instance = form.save(commit=False)
        if instance.peso and instance.talla:
            instance.imc = round(instance.peso / ((instance.talla / 100) ** 2), 2)
        instance.updated_by = self.request.user  # Auditoría
        instance.save()
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy(self.cancel_url_alt if 'from' in self.request.GET else self.cancel_url_default)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['cancel_url'] = self.cancel_url_alt if 'from' in self.request.GET else self.cancel_url_default
        return context

class EliminarTrabajadorView(LoginRequiredMixin, View):
    model = Trabajador
    template_name = 'eliminar_trabajador.html'
    pk_url_kwarg = 'trabajador_id'

    def get_object(self):
        return get_object_or_404(Trabajador, pk=self.kwargs[self.pk_url_kwarg])

    def get(self, request, *args, **kwargs):
        self.object = self.get_object()
        context = self.get_context_data()
        return render(request, self.template_name, context)

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        self.object.deleted_by = request.user  # Auditoría
        self.object.is_deleted = True
        self.object.save()
        messages.success(self.request, '[trabajadores] Trabajador eliminado correctamente.')
        return redirect(self.get_success_url())

    def get_success_url(self):
        return reverse_lazy('historial') if 'from' in self.request.GET else reverse_lazy('registro_trabajadores')

    def get_context_data(self, **kwargs):
        # Retornar el NOMBRE de la URL, no una ruta
        cancel_url_name = 'historial' if 'from' in self.request.GET else 'registro_trabajadores'
        return {
            'cancel_url': cancel_url_name,
            'trabajador': self.get_object()
        }

#-------------------
# Views Seguimientos
#-------------------

class RegistrarSeguimientoNinoView(FiltroCAIMixin, InitialFromModelMixin, RevisionCreateView):
    model = SeguimientoTrimestral
    form_class = SeguimientoTrimestralForm
    template_name = 'registrar_seguimiento.html'
    success_url = reverse_lazy('lista_seguimientos')
    comment = "Registro de seguimiento trimestral de niño"
    success_message = "[seguimiento] Seguimiento registrado correctamente."
    related_model = 'paciente'
    related_model_class = Paciente
    field_map = {
        'paciente': 'id',
        'peso': 'peso',
        'talla': 'talla',
        'imc': 'imc',
    }

    def get_form_kwargs(self):
        """Añade el usuario actual a los kwargs del formulario"""
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def form_valid(self, form):
        """Procesa el formulario cuando es válido"""
        paciente = form.cleaned_data.get('paciente')

        # Validar permisos del usuario
        if self.request.user.is_nutriologo and paciente.cai != self.request.user.cai:
            logger.warning(f"Intento de acceso no autorizado al paciente {paciente.id} por el usuario {self.request.user.id}")
            return HttpResponseForbidden("No tienes permiso para registrar seguimiento de este paciente.")

        # Guardar el seguimiento
        seguimiento = form.save(commit=False)
        seguimiento.created_by = self.request.user
        seguimiento.save()

        # Registrar acción
        logger.info(f"Seguimiento trimestral registrado para el paciente {paciente.id} por el usuario {self.request.user.id}")
        messages.success(self.request, self.success_message)
        
        return super().form_valid(form)

    def form_invalid(self, form):
        """Maneja los casos cuando el formulario es inválido"""
        logger.warning(f"Formulario inválido: {form.errors}")
        messages.error(self.request, "Por favor corrija los errores en el formulario.")
        return super().form_invalid(form)

    def get_context_data(self, **kwargs):
        """Añade datos adicionales al contexto del template"""
        context = super().get_context_data(**kwargs)
        paciente_id = self.request.GET.get("paciente_id") or self.request.POST.get("paciente")
        
        if paciente_id:
            try:
                paciente = get_object_or_404(Paciente, id=paciente_id)
                context.update({
                    "paciente_nombre": paciente.nombre,
                    "paciente_nacimiento": paciente.fecha_nacimiento.strftime('%Y-%m-%d'),
                    "paciente_sexo": paciente.sexo,
                    "paciente_edad": paciente.edad_detallada,
                })
            except Exception as e:
                logger.error(f"Error al obtener datos del paciente: {str(e)}")
                messages.error(self.request, "Error al cargar los datos del paciente.")
        
        return context

def calcular_edad_api(request):
    """API para calcular la edad basada en fechas"""
    nacimiento = request.GET.get('nacimiento')
    valoracion = request.GET.get('valoracion')

    logger.info(f"Calculando edad: nacimiento={nacimiento}, valoracion={valoracion}")

    try:
        if not nacimiento or not valoracion:
            return JsonResponse({'error': 'Fechas incompletas'}, status=400)

        nacimiento_dt = datetime.strptime(nacimiento, "%Y-%m-%d").date()
        valoracion_dt = datetime.strptime(valoracion, "%Y-%m-%d").date()

        if valoracion_dt < nacimiento_dt:
            return JsonResponse({'error': 'La fecha de valoración no puede ser anterior al nacimiento'}, status=400)

        diff = relativedelta(valoracion_dt, nacimiento_dt)
        edad_meses = diff.years * 12 + diff.months
        edad_anios = diff.years
        edad_decimal = round(edad_meses / 12, 2)
        edad_texto = f"{edad_anios} años, {diff.months} meses"

        return JsonResponse({
            'edad_texto': edad_texto,
            'edad_decimal': edad_decimal,
            'edad_meses': edad_meses
        })

    except ValueError as e:
        logger.error(f"Error en formato de fecha: {str(e)}")
        return JsonResponse({'error': 'Formato de fecha inválido. Use YYYY-MM-DD'}, status=400)
    except Exception as e:
        logger.error(f"Error al calcular edad: {str(e)}")
        return JsonResponse({'error': f'Error al calcular edad: {str(e)}'}, status=500)

def calcular_indicadores_api(request):
    """API para calcular indicadores nutricionales"""
    try:
        peso = float(request.GET.get('peso', 0))
        talla = float(request.GET.get('talla', 0))
        edad = float(request.GET.get('edad', 0))  # en años
        sexo = request.GET.get('sexo', 'F').upper()

        if peso <= 0 or talla <= 0 or edad <= 0:
            return JsonResponse({'error': 'Peso, talla y edad deben ser valores positivos'}, status=400)

        if sexo not in ['M', 'F']:
            return JsonResponse({'error': 'Sexo debe ser M o F'}, status=400)

        # Calcular IMC
        talla_m = talla / 100
        imc = round(peso / (talla_m ** 2), 2)

        # Calcular desviaciones estándar más cercanas
        peso_talla = obtener_desviacion_oms_peso_talla(sexo, talla, peso)
        peso_edad = obtener_desviacion_oms_peso_edad(sexo, edad, peso)
        talla_edad = obtener_desviacion_oms_talla_edad(sexo, edad, talla)

        # Diagnóstico basado en IMC y edad
        if edad < 5:
            if imc < 14:
                dx = "Desnutrición"
            elif imc < 17:
                dx = "Normal"
            elif imc < 19:
                dx = "Sobrepeso"
            else:
                dx = "Obesidad"
        else:
            if imc < 18.5:
                dx = "Bajo peso"
            elif imc < 25:
                dx = "Normal"
            elif imc < 30:
                dx = "Sobrepeso"
            elif imc < 35:
                dx = "Obesidad I"
            elif imc < 40:
                dx = "Obesidad II"
            else:
                dx = "Obesidad III"

        return JsonResponse({
            'peso_talla': peso_talla,
            'peso_edad': peso_edad,
            'talla_edad': talla_edad,
            'imc': imc,
            'dx': dx,
            'status': 'success'
        })

    except ValueError as e:
        return JsonResponse({'error': f'Valores inválidos: {str(e)}'}, status=400)
    except Exception as e:
        return JsonResponse({'error': f'Error interno: {str(e)}'}, status=500)


class RegistrarSeguimientoTrabajadorView(FiltroCAIMixin, InitialFromModelMixin, RevisionCreateView):
    model = SeguimientoTrabajador
    form_class = SeguimientoTrabajadorForm
    template_name = 'registrar_seguimiento_trabajador.html'
    success_url = reverse_lazy('lista_seguimientos')
    comment = "Registro de seguimiento de trabajador"
    success_message = "[seguimiento] Seguimiento registrado."
    related_model = 'trabajador'
    related_model_class = Trabajador
    field_map = {
        'trabajador': 'id',
        'peso': 'peso',
        'talla': 'talla',
        'imc': 'imc',
    }

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def form_valid(self, form):
        trabajador = form.cleaned_data.get('trabajador')

        # Restricción por CAI
        if self.request.user.is_nutriologo and trabajador.cai != self.request.user.cai:
            return HttpResponseForbidden("No tienes permiso para registrar seguimiento de este trabajador.")

        instance = form.save(commit=False)
        instance.edad = form.cleaned_data.get('edad')
        instance.imc = form.cleaned_data.get('imc')
        instance.created_by = self.request.user  # Auditoría

        with create_revision():
            instance.save()
            set_user(self.request.user)
            set_comment(self.comment)

            # Actualizar datos del trabajador
            trabajador.peso = instance.peso
            trabajador.talla = instance.talla
            trabajador.imc = instance.imc
            trabajador.circunferencia_abdominal = instance.circunferencia_abdominal
            trabajador.save()

        self.object = instance
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        trabajador_id = (
            self.request.GET.get('trabajador_id') or
            self.get_initial().get('trabajador') or
            self.request.POST.get('trabajador')
        )

        context['trabajador_nombre'] = ''
        if trabajador_id:
            try:
                trabajador = Trabajador.objects.get(id=trabajador_id)
                context['trabajador_nombre'] = trabajador.nombre
            except Trabajador.DoesNotExist:
                pass

        return context


class ListaSeguimientosGeneralView(FiltroCAIMixin, LoginRequiredMixin, View):
    model = SeguimientoTrimestral  # Necesario para que el mixin sepa qué modelo usar

    def get(self, request):
        q_nino = request.GET.get('q_nino', '')
        q_trabajador = request.GET.get('q_trabajador', '')

        # Base de datos inicial
        seguimientos_ninos = SeguimientoTrimestral.objects.select_related('paciente')
        seguimientos_trabajadores = SeguimientoTrabajador.objects.select_related('trabajador')

        # Filtro por CAI si es nutriólogo
        if request.user.is_nutriologo and request.user.cai:
            seguimientos_ninos = seguimientos_ninos.filter(paciente__cai=request.user.cai)
            seguimientos_trabajadores = seguimientos_trabajadores.filter(trabajador__cai=request.user.cai)

        # Filtros de búsqueda adicionales
        if q_nino:
            seguimientos_ninos = seguimientos_ninos.filter(paciente__nombre__icontains=q_nino)
        if q_trabajador:
            seguimientos_trabajadores = seguimientos_trabajadores.filter(trabajador__nombre__icontains=q_trabajador)

        # Renderiza el template con los seguimientos filtrados
        return render(request, 'seguimientos_general.html', {
            'seguimientos_ninos': seguimientos_ninos,
            'seguimientos_trabajadores': seguimientos_trabajadores,
            'q_nino': q_nino,
            'q_trabajador': q_trabajador
        })

class SeguimientosNinoView(FiltroCAIMixin, LoginRequiredMixin, View):
    def get(self, request, nino_id):
        paciente = get_object_or_404(Paciente, id=nino_id)

        if request.user.is_nutriologo and paciente.cai != request.user.cai:
            raise Http404("No autorizado")

        seguimientos = SeguimientoTrimestral.objects.filter(
            paciente=paciente,
            is_deleted=False
        ).order_by('fecha_valoracion')

        datos = {
            'fechas': [s.fecha_valoracion.strftime('%d/%m/%Y') for s in seguimientos if s.fecha_valoracion],
            'pesos': [float(s.peso) for s in seguimientos if s.peso is not None],
            'tallas': [float(s.talla) for s in seguimientos if s.talla is not None],
            'imcs': [float(s.imc) for s in seguimientos if s.imc is not None],
        }

        return render(request, 'seguimientos_nino.html', {
            'nino': paciente,
            'seguimientos': seguimientos,
            'datos_json': json.dumps(datos),
        })

class SeguimientosTrabajadorView(LoginRequiredMixin, View):
    def get(self, request, trabajador_id):
        trabajador = get_object_or_404(Trabajador, id=trabajador_id)

        if request.user.is_nutriologo and trabajador.cai != request.user.cai:
            return HttpResponseForbidden("No tienes acceso a este trabajador.")

        seguimientos = SeguimientoTrabajador.objects.filter(trabajador=trabajador).order_by('fecha_valoracion')

        datos = {
            'fechas': [DateFormat(s.fecha_valoracion).format('d/m/Y') for s in seguimientos if s.fecha_valoracion],
            'pesos': [float(s.peso) for s in seguimientos if s.peso is not None],
            'tallas': [float(s.talla) for s in seguimientos if s.talla is not None],
            'imcs': [float(s.imc) for s in seguimientos if s.imc is not None],
        }

        return render(request, 'seguimientos_trabajador.html', {
            'trabajador': trabajador,
            'seguimientos': seguimientos,
            'datos_json': json.dumps(datos),
        })
    
class EditarSeguimientoNinoView(FiltroCAIMixin, LoginRequiredMixin, View):
    def get(self, request, id):
        seguimiento = get_object_or_404(SeguimientoTrimestral, id=id)
        if request.user.is_nutriologo and seguimiento.paciente.cai != request.user.cai:
            return HttpResponseForbidden("No tienes permiso para editar este seguimiento.")

        form = SeguimientoTrimestralForm(instance=seguimiento, user=request.user)
        return render(request, 'editar_seguimiento_nino.html', {
            'form': form,
            'seguimiento': seguimiento
        })

    def post(self, request, id):
        seguimiento = get_object_or_404(SeguimientoTrimestral, id=id)
        if request.user.is_nutriologo and seguimiento.paciente.cai != request.user.cai:
            return HttpResponseForbidden("No tienes permiso para editar este seguimiento.")

        form = SeguimientoTrimestralForm(request.POST, instance=seguimiento, user=request.user)
        if form.is_valid():
            instance = form.save(commit=False)
            instance.updated_by = request.user # Auditoría
            with create_revision():
                instance.save()
                set_user(request.user)
                set_comment("Edición de seguimiento de niño.")
            return redirect('seguimientos_nino', nino_id=seguimiento.paciente.id)
        return render(request, 'editar_seguimiento_nino.html', {
            'form': form,
            'seguimiento': seguimiento
        })

class EditarSeguimientoTrabajadorView(LoginRequiredMixin, View):
    def get(self, request, id):
        seguimiento = get_object_or_404(SeguimientoTrabajador, id=id)

        if request.user.is_nutriologo and seguimiento.trabajador.cai != request.user.cai:
            return HttpResponseForbidden("No tienes permiso para editar este seguimiento.")

        form = SeguimientoTrabajadorForm(instance=seguimiento, user=request.user)
        return render(request, 'editar_seguimiento_trabajador.html', {'form': form, 'seguimiento': seguimiento})

    def post(self, request, id):
        seguimiento = get_object_or_404(SeguimientoTrabajador, id=id)

        if request.user.is_nutriologo and seguimiento.trabajador.cai != request.user.cai:
            return HttpResponseForbidden("No tienes permiso para editar este seguimiento.")

        form = SeguimientoTrabajadorForm(request.POST, instance=seguimiento, user=request.user)
        if form.is_valid():
            seguimiento_actualizado = form.save(commit=False)
            seguimiento_actualizado.updated_by = request.user # Auditoría

            trabajador = seguimiento_actualizado.trabajador
            trabajador.peso = seguimiento_actualizado.peso
            trabajador.talla = seguimiento_actualizado.talla
            trabajador.imc = seguimiento_actualizado.imc
            trabajador.circunferencia_abdominal = seguimiento_actualizado.circunferencia_abdominal

            with create_revision():
                seguimiento_actualizado.save()
                trabajador.save()
                set_user(request.user)
                set_comment("Edición de seguimiento y actualización de datos del trabajador.")

            return redirect('seguimientos_trabajador', trabajador_id=trabajador.id)

        return render(request, 'editar_seguimiento_trabajador.html', {'form': form, 'seguimiento': seguimiento})

class ListaSeguimientosView(LoginRequiredMixin, View):
    def get(self, request):
        # Si es AJAX, responder con JSON
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return self.handle_ajax(request)

        return render(request, 'seguimientos_general.html', {
            'initial_data': {
                'ninos': self.get_seguimientos_ninos(request),
                'trabajadores': self.get_seguimientos_trabajadores(request)
            }
        })

    def handle_ajax(self, request):
        term = request.GET.get('term', '')
        tipo = request.GET.get('type', '')

        if tipo == 'nino':
            data = self.get_seguimientos_ninos(request, search_term=term)
        elif tipo == 'trabajador':
            data = self.get_seguimientos_trabajadores(request, search_term=term)
        else:
            return JsonResponse({'error': 'Tipo inválido'}, status=400)

        return JsonResponse({'resultados': data, 'has_next': False})

    def get_seguimientos_ninos(self, request, search_term=None):
        qs = SeguimientoTrimestral.objects.select_related('paciente') \
            .filter(is_deleted=False, paciente__is_deleted=False) \
            .order_by('paciente__id', '-fecha_valoracion')

        if request.user.is_nutriologo and request.user.cai:
            qs = qs.filter(paciente__cai=request.user.cai)

        if search_term:
            qs = qs.filter(paciente__nombre__icontains=search_term)

        vistos = set()
        resultados = []

        for s in qs:
            if s.paciente.id not in vistos:
                vistos.add(s.paciente.id)
                resultados.append({
                    'id': s.id,
                    'paciente_id': s.paciente.id,
                    'nombre': s.paciente.nombre,
                    'fecha': s.fecha_valoracion.strftime('%d/%m/%Y') if s.fecha_valoracion else '',
                    'edad': calcular_edad_detallada(s.paciente.fecha_nacimiento, s.fecha_valoracion),
                    'peso': s.peso,
                    'talla': s.talla,
                    'imc': f"{s.imc:.2f}" if s.imc else '',
                    'dx': s.dx or '',
                    'url': reverse('seguimientos_nino', kwargs={'nino_id': s.paciente.id})  # ✅ CORREGIDO
                })

        return resultados

    def get_seguimientos_trabajadores(self, request, search_term=None):
        qs = SeguimientoTrabajador.objects.select_related('trabajador') \
            .filter(is_deleted=False, trabajador__is_deleted=False) \
            .order_by('trabajador__id', '-fecha_valoracion')

        if request.user.is_nutriologo and request.user.cai:
            qs = qs.filter(trabajador__cai=request.user.cai)

        if search_term:
            qs = qs.filter(trabajador__nombre__icontains=search_term)

        vistos = set()
        resultados = []

        for s in qs:
            if s.trabajador.id not in vistos:
                vistos.add(s.trabajador.id)
                resultados.append({
                    'id': s.id,
                    'trabajador_id': s.trabajador.id,
                    'nombre': s.trabajador.nombre,
                    'fecha': s.fecha_valoracion.strftime('%d/%m/%Y') if s.fecha_valoracion else '',
                    'edad': s.trabajador.edad or '—',
                    'peso': s.peso,
                    'talla': s.talla,
                    'imc': f"{s.imc:.2f}" if s.imc else '',
                    'dx': s.dx or '',
                    'url': reverse('seguimientos_trabajador', kwargs={'trabajador_id': s.trabajador.id})
                })

        return resultados

class EliminarSeguimientoNinoView(LoginRequiredMixin, View):
    def get_object(self, pk):
        return get_object_or_404(SeguimientoTrimestral, pk=pk)

    def get(self, request, pk):
        seguimiento = self.get_object(pk)

        if seguimiento.is_deleted:
            messages.warning(request, "Este seguimiento ya fue eliminado.")
            return redirect('seguimientos_general')

        if request.user.is_nutriologo and seguimiento.paciente.cai != request.user.cai:
            return HttpResponseForbidden("No tienes permiso para eliminar este seguimiento.")

        return render(request, 'eliminar_seguimiento.html', {'seguimiento': seguimiento})

    def post(self, request, pk):
        seguimiento = self.get_object(pk)

        if seguimiento.is_deleted:
            messages.warning(request, "Este seguimiento ya fue eliminado.")
            return redirect('seguimientos_general')

        if request.user.is_nutriologo and seguimiento.paciente.cai != request.user.cai:
            return HttpResponseForbidden("No tienes permiso para eliminar este seguimiento.")

        seguimiento.deleted_by = request.user
        seguimiento.is_deleted = True
        seguimiento.save()
        messages.success(request, "Seguimiento eliminado correctamente.")
        return redirect('seguimientos_general')
    
class EliminarSeguimientoTrabajadorView(LoginRequiredMixin, View):
    def get_object(self, pk):
        return get_object_or_404(SeguimientoTrabajador, pk=pk)

    def get(self, request, pk):
        seguimiento = self.get_object(pk)

        if seguimiento.is_deleted:
            messages.warning(request, "Este seguimiento ya fue eliminado.")
            return redirect('seguimientos_general')

        if request.user.is_nutriologo and seguimiento.trabajador.cai != request.user.cai:
            return HttpResponseForbidden("No tienes permiso para eliminar este seguimiento.")

        return render(request, 'eliminar_seguimiento.html', {'seguimiento': seguimiento})

    def post(self, request, pk):
        seguimiento = self.get_object(pk)

        if seguimiento.is_deleted:
            messages.warning(request, "Este seguimiento ya fue eliminado.")
            return redirect('seguimientos_general')

        if request.user.is_nutriologo and seguimiento.trabajador.cai != request.user.cai:
            return HttpResponseForbidden("No tienes permiso para eliminar este seguimiento.")

        seguimiento.deleted_by = request.user  # Auditoría
        seguimiento.is_deleted = True
        seguimiento.save()
        messages.success(request, "Seguimiento eliminado correctamente.")
        return redirect('seguimientos_general')

#-------------------
# Views Historial
#-------------------

class HistorialView(LoginRequiredMixin, TemplateView):
    template_name = 'historial.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        q_nino = self.request.GET.get('q_nino', '')
        q_trabajador = self.request.GET.get('q_trabajador', '')

        user = self.request.user

        pacientes = Paciente.objects.filter(is_deleted=False)
        trabajadores = Trabajador.objects.filter(is_deleted=False)

        # Filtro por CAI si es nutriólogo
        if user.is_nutriologo and user.cai:
            pacientes = pacientes.filter(cai=user.cai)
            trabajadores = trabajadores.filter(cai=user.cai)

        context.update({
            'request': self.request,
            'pacientes': pacientes,
            'trabajadores': trabajadores,
            'q_nino': q_nino,
            'q_trabajador': q_trabajador,
            'resultados_ninos': pacientes.filter(nombre__icontains=q_nino) if q_nino else [],
            'resultados_trabajadores': trabajadores.filter(nombre__icontains=q_trabajador) if q_trabajador else [],
            'puede_editar': user.is_admin or user.is_jefe_departamento or user.is_nutriologo,
            'puede_eliminar': user.is_admin or user.is_jefe_departamento,
        })
        return context

class HistorialNinoCambiosView(LoginRequiredMixin, View):
    def get(self, request, nino_id):
        nino = get_object_or_404(Paciente, id=nino_id)
        versiones = Version.objects.get_for_object(nino)
        return render(request, 'historial_nino_cambios.html', {'nino': nino, 'versiones': versiones})
    
#-------------------
# Views Reportes
#-------------------

class ReportesView(LoginRequiredMixin, FiltroCAIMixin, TemplateView):
    template_name = 'reportes.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.request.user

        # Verificar si es nutriólogo
        es_nutriologo = hasattr(user, 'cai') and not (user.is_admin or user.is_jefe_departamento)
        cai_usuario = user.cai if es_nutriologo else ''

        # Subqueries para verificar existencia de seguimientos
        subq_pacientes = SeguimientoTrimestral.objects.filter(
            paciente=OuterRef('pk'),
            is_deleted=False
        )
        subq_trabajadores = SeguimientoTrabajador.objects.filter(
            trabajador=OuterRef('pk'),
            is_deleted=False
        )

        # Filtrar según tipo de usuario
        if user.is_admin or user.is_jefe_departamento:
            pacientes = Paciente.objects.filter(is_deleted=False)
            trabajadores = Trabajador.objects.filter(is_deleted=False)
        else:
            pacientes = Paciente.objects.filter(is_deleted=False, cai=user.cai)
            trabajadores = Trabajador.objects.filter(is_deleted=False, cai=user.cai)

        # Anotar si tienen seguimiento
        pacientes = pacientes.annotate(tiene_seguimiento=Exists(subq_pacientes)).order_by('nombre')
        trabajadores = trabajadores.annotate(tiene_seguimiento=Exists(subq_trabajadores)).order_by('nombre')

        # Asignar edad actualizada si hay seguimiento
        for paciente in pacientes:
            seguimiento = SeguimientoTrimestral.objects.filter(
                paciente=paciente,
                is_deleted=False
            ).order_by('-fecha_valoracion').first()

            paciente.edad_seguimiento = seguimiento.edad if seguimiento and seguimiento.edad else paciente.edad_detallada
            paciente.tiene_seguimiento = seguimiento is not None

        for trabajador in trabajadores:
            seguimiento = SeguimientoTrabajador.objects.filter(
                trabajador=trabajador,
                is_deleted=False
            ).order_by('-fecha_valoracion').first()

            trabajador.tiene_seguimiento = seguimiento is not None

        # Paginación
        paginator_pacientes = Paginator(pacientes, 10)
        paginator_trabajadores = Paginator(trabajadores, 10)

        context['ninos_page'] = paginator_pacientes.get_page(self.request.GET.get('page_pacientes', 1))
        context['trabajadores_page'] = paginator_trabajadores.get_page(self.request.GET.get('page_trabajadores', 1))
        context['es_nutriologo'] = es_nutriologo
        context['mi_cai'] = cai_usuario

        return context

class ReporteBaseView(LoginRequiredMixin, View):
    """Vista base para reportes individuales"""
    template_name = None
    model = None
    seguimiento_model = None
    id_kwarg = 'pk'

    def get(self, request, *args, **kwargs):
        obj_id = kwargs.get(self.id_kwarg)
        try:
            obj = self.model.objects.get(pk=obj_id)
            print(f"Objeto encontrado: {obj}")  # Debug

            # Determinar el campo de relación
            filter_field = 'paciente' if isinstance(obj, Paciente) else 'trabajador'
            seguimientos = self.seguimiento_model.objects.filter(
                is_deleted=False,
                **{filter_field: obj}
            ).order_by('fecha_valoracion')

            # Preparar datos para gráficos
            datos = {
                'fechas': [s.fecha_valoracion.strftime('%Y-%m-%d') for s in seguimientos],
                'pesos': [float(s.peso) if s.peso else None for s in seguimientos],
                'imcs': [float(s.imc) if s.imc else None for s in seguimientos],
            }

            # Añadir tallas solo para pacientes
            if hasattr(self.seguimiento_model, 'talla'):
                datos['tallas'] = [float(s.talla) if s.talla else None for s in seguimientos]

            context = {
                'paciente' if isinstance(obj, Paciente) else 'trabajador': obj,
                'seguimientos': seguimientos,
                'datos_json': json.dumps(datos),
                'debug_mode': True  # Para mostrar info de debug en template
            }

            return render(request, self.template_name, context)

        except self.model.DoesNotExist:
            raise Http404("El registro solicitado no existe")
        except Exception as e:
            print(f"Error en ReporteBaseView: {str(e)}")  # Debug
            raise

class ReportePacienteView(View):
    template_name = 'reporte_paciente.html'
    model = Paciente
    seguimiento_model = SeguimientoTrimestral
    id_kwarg = 'pk'

    def get_datos_oms(self, modelo, campo_x, sexo):
        filas = modelo.objects.filter(sexo=sexo).order_by(campo_x)
        etiquetas = [getattr(f, campo_x) for f in filas]
        curvas = {
            '-3 SD': [f.sd_m3 for f in filas],
            '-2 SD': [f.sd_m2 for f in filas],
            '-1 SD': [f.sd_m1 for f in filas],
            'Mediana': [f.mediana for f in filas],
            '+1 SD': [f.sd_1 for f in filas],
            '+2 SD': [f.sd_2 for f in filas],
            '+3 SD': [f.sd_3 for f in filas],
        }
        return etiquetas, curvas

    def get(self, request, *args, **kwargs):
        obj_id = kwargs.get(self.id_kwarg)
        try:
            paciente = self.model.objects.get(pk=obj_id, is_deleted=False)
            seguimientos = self.seguimiento_model.objects.filter(
                paciente=paciente,
                is_deleted=False
            ).order_by('fecha_valoracion')

            sexo = paciente.sexo.upper()

            # Calcular edad detallada a partir del último seguimiento
            if seguimientos.exists() and paciente.fecha_nacimiento:
                fecha_base = seguimientos.last().fecha_valoracion
                edad_rd = relativedelta(fecha_base, paciente.fecha_nacimiento)
                edad_actualizada = f"{edad_rd.years} años, {edad_rd.months} meses"
            else:
                edad_actualizada = "Edad no disponible"

            # Datos del paciente por seguimiento
            fechas = []
            edad_meses = []
            talla_cm = []

            peso_edad = []
            talla_edad = []
            peso_talla = []

            for s in seguimientos:
                fechas.append(s.fecha_valoracion.strftime('%Y-%m-%d'))

                edad_rd = relativedelta(s.fecha_valoracion, paciente.fecha_nacimiento)
                edad_num = round((edad_rd.years * 12 + edad_rd.months), 2)
                edad_meses.append(edad_num)

                peso = float(s.peso) if s.peso else None
                talla = float(s.talla) if s.talla else None
                if talla: talla_cm.append(round(talla, 1))
                else: talla_cm.append(None)

                peso_edad.append(peso)
                talla_edad.append(talla)
                peso_talla.append(peso)

            # JSON para Chart.js
            datos = {
                'fechas': fechas,
                'edad_meses': edad_meses,
                'talla_cm': talla_cm,
                'peso_edad': peso_edad,
                'talla_edad': talla_edad,
                'peso_talla': peso_talla,
            }

            etiquetas_pe, curvas_pe = self.get_datos_oms(OmsPesoEdad, 'meses', sexo)
            etiquetas_te, curvas_te = self.get_datos_oms(OmsTallaEdad, 'meses', sexo)
            etiquetas_pt, curvas_pt = self.get_datos_oms(OmsPesoTalla, 'talla_cm', sexo)

            context = {
                'paciente': paciente,
                'seguimientos': seguimientos,
                'edad_actualizada': edad_actualizada,
                'datos_json': json.dumps(datos),
                'oms_data': json.dumps({
                    'peso_edad': {'etiquetas': etiquetas_pe, 'curvas': curvas_pe},
                    'talla_edad': {'etiquetas': etiquetas_te, 'curvas': curvas_te},
                    'peso_talla': {'etiquetas': etiquetas_pt, 'curvas': curvas_pt},
                }),
                'debug_mode': settings.DEBUG,
            }
            return render(request, self.template_name, context)

        except self.model.DoesNotExist:
            raise Http404("El paciente no existe")
        except Exception as e:
            print(f" Error en ReportePacienteView: {str(e)}")
            raise

class ReporteTrabajadorView(LoginRequiredMixin, DetailView):
    model = Trabajador
    template_name = 'reporte_trabajador.html'
    context_object_name = 'trabajador'

    def get_object(self):
        return get_object_or_404(Trabajador, pk=self.kwargs['pk'], is_deleted=False)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        trabajador = self.object
        trabajador.refresh_from_db()

        seguimientos = SeguimientoTrabajador.objects.filter(
            trabajador=trabajador,
            is_deleted=False
        ).order_by('fecha_valoracion')

        ultimo_seguimiento = seguimientos.last() if seguimientos.exists() else None

        # Cálculo de valores ideales:
        peso_ideal = None
        imc_ideal = None
        circ_abd_ideal = None

        if trabajador.talla:  # Suponemos talla en cm
            talla_m = trabajador.talla / 100  # convertir a metros

            # IMC ideal: promedio 22 (puedes ajustar si lo deseas)
            imc_ideal = 22  
            peso_ideal = round(imc_ideal * (talla_m ** 2), 2)

            # Circunferencia abdominal recomendada OMS:
            if trabajador.sexo == 'M':
                circ_abd_ideal = 'Menos de 94 cm'
            elif trabajador.sexo == 'F':
                circ_abd_ideal = 'Menos de 80 cm'
            else:
                circ_abd_ideal = 'No especificado'

        # Preparar datos para gráficas
        datos = {
            'fechas': [s.fecha_valoracion.strftime('%Y-%m-%d') for s in seguimientos],
            'pesos': [float(s.peso) if s.peso else None for s in seguimientos],
            'imcs': [float(s.imc) if s.imc else None for s in seguimientos],
            'circ_abdominal': [
                float(s.circunferencia_abdominal) if s.circunferencia_abdominal is not None else None
                for s in seguimientos
            ],
            'peso_ideal': peso_ideal,
            'imc_ideal': imc_ideal,
            'circ_abd_ideal': circ_abd_ideal,
        }

        context.update({
            'seguimientos': seguimientos,
            'ultimo_seguimiento': ultimo_seguimiento,
            'datos_json': json.dumps(datos),
            'peso_ideal': peso_ideal,
            'imc_ideal': imc_ideal,
            'circ_abd_ideal': circ_abd_ideal,
            'debug_mode': settings.DEBUG
        })

        return context
 
#-------------------
# Extras
#-------------------

def remover_tildes(texto):
    return ''.join(c for c in unicodedata.normalize('NFD', texto) if unicodedata.category(c) != 'Mn')
    
def buscar_ninos_ajax(request):
    term = request.GET.get('term', '').strip()
    cai_filter = request.GET.get('cai', '').strip()
    edad_filter = request.GET.get('edad', '').strip()
    page = int(request.GET.get('page', 1))

    # Filtrar por nombre y que no estén eliminados
    ninos_qs = Paciente.objects.filter(nombre__icontains=term, is_deleted=False).order_by('-id')

    # Filtrar por CAI
    if cai_filter:
        ninos = [n for n in ninos_qs if remover_tildes(n.cai.upper()) == cai_filter.upper()]
    elif request.user.is_nutriologo and request.user.cai:
        ninos = [n for n in ninos_qs if remover_tildes(n.cai.upper()) == remover_tildes(request.user.cai.upper())]
    else:
        ninos = list(ninos_qs)

    # Filtro por edad
    if edad_filter:
        hoy = date.today()
        if edad_filter == "0-2":
            min_date = hoy.replace(year=hoy.year - 2)
            max_date = hoy
        elif edad_filter == "3-5":
            min_date = hoy.replace(year=hoy.year - 5)
            max_date = hoy.replace(year=hoy.year - 3)
        elif edad_filter == "6-12":
            min_date = hoy.replace(year=hoy.year - 12)
            max_date = hoy.replace(year=hoy.year - 6)
        elif edad_filter == "13-18":
            min_date = hoy.replace(year=hoy.year - 18)
            max_date = hoy.replace(year=hoy.year - 13)
        else:
            min_date = max_date = None

        if min_date and max_date:
            ninos = [n for n in ninos if min_date <= n.fecha_nacimiento <= max_date]

    # Paginación
    paginator = Paginator(ninos, 10)
    pagina = paginator.get_page(page)

    resultados = []
    for n in pagina:
        # Buscar último seguimiento
        seguimiento = SeguimientoTrimestral.objects.filter(paciente=n, is_deleted=False).order_by('-fecha_valoracion').first()

        if seguimiento:
            edad_rd = relativedelta(seguimiento.fecha_valoracion, n.fecha_nacimiento)
            edad_texto = f"{edad_rd.years} años, {edad_rd.months} meses"
            edad_anios = edad_rd.years
            tiene_seguimiento = True
        else:
            edad_rd = relativedelta(date.today(), n.fecha_nacimiento)
            edad_texto = f"{edad_rd.years} años, {edad_rd.months} meses"
            edad_anios = edad_rd.years
            tiene_seguimiento = False

        resultados.append({
            'id': n.id,
            'nombre': n.nombre,
            'edad': edad_texto,
            'edad_anios': edad_anios,
            'curp': n.curp,
            'grado': n.grado,
            'grupo': n.grupo,
            'cai': n.cai,
            'tiene_seguimiento': tiene_seguimiento
        })

    return JsonResponse({
        'resultados': resultados,
        'has_next': pagina.has_next()
    })

def buscar_trabajadores_ajax(request):
    term = request.GET.get('term', '')
    cai_filter = request.GET.get('cai', '')
    page = request.GET.get('page', 1)

    trabajadores = Trabajador.objects.filter(nombre__icontains=term, is_deleted=False)

    if cai_filter:
        trabajadores = trabajadores.filter(cai=cai_filter)
    elif request.user.is_nutriologo and request.user.cai:
        trabajadores = trabajadores.filter(cai=request.user.cai)

    trabajadores = trabajadores.order_by('nombre')
    paginator = Paginator(trabajadores, 10)
    pagina = paginator.get_page(page)

    resultados = [
        {
            'id': t.id,
            'nombre': t.nombre,
            'curp': t.curp,
            'cargo': t.cargo,
            'departamento': t.departamento,
            'cai': str(t.cai),
        } for t in pagina
    ]

    return JsonResponse({
        'resultados': resultados,
        'has_next': pagina.has_next()
    })

def calcular_edad_detallada(nacimiento, valoracion):
    if not nacimiento or not valoracion:
        return "0 años, 0 meses"
    diferencia = relativedelta(valoracion, nacimiento)
    return f"{diferencia.years} años, {diferencia.months} meses"

def buscar_seguimientos_nino_ajax(request):
    termino = request.GET.get('term', '')

    qs = SeguimientoTrimestral.objects.select_related('paciente').filter(
        is_deleted=False,
        paciente__is_deleted=False
    ).order_by('paciente__id', '-fecha_valoracion')

    if request.user.is_nutriologo and request.user.cai:
        qs = qs.filter(paciente__cai=request.user.cai)

    if termino:
        qs = qs.filter(paciente__nombre__icontains=termino)

    # Agrupar manualmente por paciente y tomar solo el último válido
    vistos = set()
    resultados = []
    for s in qs:
        if s.paciente.id not in vistos:
            vistos.add(s.paciente.id)
            resultados.append({
                'id': s.id,
                'paciente_id': s.paciente.id,
                'nombre': s.paciente.nombre,
                'fecha': s.fecha_valoracion.strftime('%d/%m/%Y') if s.fecha_valoracion else '',
                'edad': calcular_edad_detallada(s.paciente.fecha_nacimiento, s.fecha_valoracion),
                'peso': s.peso,
                'talla': s.talla,
                'imc': s.imc,
                'dx': s.dx or ''
            })

    return JsonResponse({'resultados': resultados})

def buscar_seguimientos_trabajador_ajax(request):
    termino = request.GET.get('term', '')

    base_queryset = SeguimientoTrabajador.objects.select_related('trabajador').filter(is_deleted=False, trabajador__is_deleted=False)

    if request.user.is_nutriologo and request.user.cai:
        base_queryset = base_queryset.filter(trabajador__cai=request.user.cai)

    latest_dates = base_queryset.values('trabajador').annotate(
        max_fecha=Max('fecha_valoracion')
    ).values('trabajador', 'max_fecha')

    queryset = base_queryset.filter(
        trabajador__nombre__icontains=termino,
        fecha_valoracion__in=Subquery(
            latest_dates.filter(trabajador=OuterRef('trabajador')).values('max_fecha')
        )
    ).order_by('trabajador__nombre')

    data = [{
        'id': s.id,
        'trabajador_id': s.trabajador.id,
        'nombre': s.trabajador.nombre,
        'fecha': s.fecha_valoracion.strftime('%d/%m/%Y') if s.fecha_valoracion else '',
        'edad': s.edad,
        'peso': s.peso,
        'talla': s.talla,
        'imc': s.imc,
        'dx': s.dx or ''
    } for s in queryset]

    return JsonResponse({'resultados': data})

#Implemenatción de reportes con weasyprint

def generar_grafica_oms_completa(seguimientos, tipo="peso", sexo="M"):
    import matplotlib.pyplot as plt
    from nutricion.models import OmsPesoEdad, OmsTallaEdad, OmsPesoTalla
    from dateutil.relativedelta import relativedelta

    if not seguimientos:
        return None

    # Seleccionar modelo y ejes
    if tipo == "peso":
        modelo = OmsPesoEdad
        campo_x = "meses"
        campo_y = "peso"
        x_label = "Edad (meses)"
        y_label = "z-score/desviación estándar"
        titulo = "Peso para la Edad según OMS"
    elif tipo == "talla":
        modelo = OmsTallaEdad
        campo_x = "meses"
        campo_y = "talla"
        x_label = "Talla (cm)"
        y_label = "z-score/desviación estándar"
        titulo = "Talla para la Edad según OMS"
    elif tipo == "peso_talla":
        modelo = OmsPesoTalla
        campo_x = "talla_cm"
        campo_y = "peso"
        x_label = "Talla (cm)"
        y_label = "z-score/desviación estándar"
        titulo = "Peso para la Talla según OMS"
    else:
        return None

    registros = modelo.objects.filter(sexo=sexo.upper()).order_by(campo_x)
    x_vals = [getattr(r, campo_x) for r in registros]
    curvas = {
        '-3 SD': [r.sd_m3 for r in registros],
        '-2 SD': [r.sd_m2 for r in registros],
        '-1 SD': [r.sd_m1 for r in registros],
        'Mediana': [r.mediana for r in registros],
        '+1 SD': [r.sd_1 for r in registros],
        '+2 SD': [r.sd_2 for r in registros],
        '+3 SD': [r.sd_3 for r in registros],
    }

    colores = {
        '-3 SD': '#800000',
        '-2 SD': '#FF6600',
        '-1 SD': '#FFD700',
        'Mediana': '#9ACD32',
        '+1 SD': '#FFD700',
        '+2 SD': '#FF6600',
        '+3 SD': '#800000',
    }

    # Puntos del paciente
    puntos_x = []
    puntos_y = []

    for s in seguimientos:
        if tipo == "peso":
            edad = relativedelta(s.fecha_valoracion, s.paciente.fecha_nacimiento)
            edad_meses = round((edad.years * 12 + edad.months), 2)
            if s.peso:
                puntos_x.append(edad_meses)
                puntos_y.append(s.peso)
        elif tipo == "talla":
            edad = relativedelta(s.fecha_valoracion, s.paciente.fecha_nacimiento)
            edad_meses = round((edad.years * 12 + edad.months), 2)
            if s.talla:
                puntos_x.append(edad_meses)
                puntos_y.append(s.talla)
        elif tipo == "peso_talla":
            if s.talla and s.peso:
                puntos_x.append(round(s.talla, 1))
                puntos_y.append(s.peso)

    # Dibujar
    plt.figure(figsize=(10, 5))
    for label, y_vals in curvas.items():
        plt.plot(x_vals, y_vals, label=label, color=colores[label], linewidth=2)

    if puntos_x:
        plt.plot(puntos_x, puntos_y, 'o-', color='black', label='Paciente')

    plt.title(titulo)
    plt.xlabel(x_label)
    plt.ylabel(y_label)
    plt.grid(True, linestyle='--', alpha=0.6)
    plt.legend()
    plt.tight_layout()

    buffer = io.BytesIO()
    plt.savefig(buffer, format='png', dpi=150)
    plt.close()
    return base64.b64encode(buffer.getvalue()).decode('utf-8')

def generar_pdf_paciente(request, paciente_id):
    paciente = get_object_or_404(Paciente, pk=paciente_id)
    seguimientos = SeguimientoTrimestral.objects.filter(paciente=paciente, is_deleted=False).order_by('fecha_valoracion')

    # Función robusta para obtener imágenes como base64
    def get_image_base64(relative_static_path):
        try:
            # Producción: buscar en STATIC_ROOT
            if hasattr(settings, 'STATIC_ROOT') and settings.STATIC_ROOT:
                abs_path = os.path.join(settings.STATIC_ROOT, relative_static_path)
                if os.path.exists(abs_path):
                    with open(abs_path, "rb") as img:
                        return base64.b64encode(img.read()).decode('utf-8')

            # Desarrollo: buscar en STATICFILES_DIRS
            for static_dir in getattr(settings, 'STATICFILES_DIRS', []):
                abs_path = os.path.join(static_dir, relative_static_path)
                if os.path.exists(abs_path):
                    with open(abs_path, "rb") as img:
                        return base64.b64encode(img.read()).decode('utf-8')

            # Fallback: intentar en la app directamente
            fallback_path = os.path.join('nutricion', 'static', relative_static_path)
            if os.path.exists(fallback_path):
                with open(fallback_path, "rb") as img:
                    return base64.b64encode(img.read()).decode('utf-8')

            print(f"[WARN] Imagen no encontrada: {relative_static_path}")
            return None
        except Exception as e:
            print(f"[ERROR] al cargar imagen {relative_static_path}: {str(e)}")
            return None

    # Cargar imágenes institucionales
    logo_base64 = get_image_base64('nutricion/images/dofyas.png')
    fondo_base64 = get_image_base64('nutricion/images/fondodif2.png')

    # Gráficas tipo OMS
    grafica_peso_oms_base64 = generar_grafica_oms_completa(seguimientos, tipo="peso", sexo=paciente.sexo)
    grafica_talla_oms_base64 = generar_grafica_oms_completa(seguimientos, tipo="talla", sexo=paciente.sexo)
    grafica_peso_talla_oms_base64 = generar_grafica_oms_completa(seguimientos, tipo="peso_talla", sexo=paciente.sexo)

    # Render HTML del reporte
    html_string = render_to_string('reporte_paciente_pdf.html', {
        'paciente': paciente,
        'seguimientos': seguimientos,
        'logo_base64': logo_base64,
        'fondo_base64': fondo_base64,
        'grafica_peso_oms_base64': grafica_peso_oms_base64,
        'grafica_talla_oms_base64': grafica_talla_oms_base64,
        'grafica_peso_talla_oms_base64': grafica_peso_talla_oms_base64,
        'pdf_mode': True
    })

    # Estilos CSS
    css_string = f'''
        @page {{
            size: A4;
            margin: 0;
            background-image: url("data:image/png;base64,{fondo_base64}");
            background-size: cover;
            background-position: center center;
            background-repeat: no-repeat;
        }}

        body {{
            font-family: Arial, sans-serif;
            margin: 0;
            padding: 0;
        }}

        .contenedor-principal {{
            background-color: rgba(255, 255, 255, 0.93);
            padding: 2cm;
            position: relative;
            z-index: 2;
        }}

        .header-institucional {{
            page-break-after: avoid;
        }}

        .seccion {{
            page-break-inside: avoid;
        }}

        .graficas-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(300px, 1fr));
            gap: 20px;
            margin-top: 20px;
        }}

        .grafica-container {{
            page-break-inside: avoid;
            margin-bottom: 30px;
        }}

        .grafica-container h4 {{
            text-align: center;
            color: #6D0000;
            margin-bottom: 10px;
            font-size: 14px;
        }}

        .logo-dif {{
            position: absolute;
            top: -1.8cm;
            left: 0.5cm;
            height: 60px;
            z-index: 100;
        }}
    '''

    # Generar PDF
    font_config = FontConfiguration()
    html = HTML(string=html_string, base_url=request.build_absolute_uri())
    css = CSS(string=css_string)
    pdf = html.write_pdf(stylesheets=[css], font_config=font_config)

    # Enviar PDF como archivo descargable
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="reporte_{paciente.nombre.replace(" ", "_")}.pdf"'
    return response

#------------ Generar PDF para trabajador-------------
def generar_pdf_trabajador(request, trabajador_id):
    trabajador = get_object_or_404(Trabajador, pk=trabajador_id)
    seguimientos = SeguimientoTrabajador.objects.filter(trabajador=trabajador).order_by('fecha_valoracion')

    # --- Función para cargar imágenes como base64 ---
    def get_image_base64(relative_static_path):
        try:
            if hasattr(settings, 'STATIC_ROOT') and settings.STATIC_ROOT:
                absolute_path = os.path.join(settings.STATIC_ROOT, relative_static_path)
                if os.path.exists(absolute_path):
                    with open(absolute_path, "rb") as img_file:
                        return base64.b64encode(img_file.read()).decode('utf-8')
            for static_dir in getattr(settings, 'STATICFILES_DIRS', []):
                absolute_path = os.path.join(static_dir, relative_static_path)
                if os.path.exists(absolute_path):
                    with open(absolute_path, "rb") as img_file:
                        return base64.b64encode(img_file.read()).decode('utf-8')
            return None
        except Exception as e:
            print(f"Error al cargar imagen {relative_static_path}: {str(e)}")
            return None

    logo_base64 = get_image_base64('nutricion/images/dofyas.png')
    fondo_base64 = get_image_base64('nutricion/images/fondodif2.png')

    plt.style.use('seaborn-v0_8')
    plt.rcParams['figure.facecolor'] = 'white'
    plt.rcParams['axes.facecolor'] = 'white'

    fechas = [s.fecha_valoracion for s in seguimientos]

    # --- Valores de referencia ---
    peso_ideal = round((trabajador.talla / 100) ** 2 * 22, 1) if trabajador.talla else None
    imc_ideal = 22
    circ_ideal = 94 if trabajador.sexo == 'M' else 80

    # --- Gráfica Peso ---
    grafica_peso_base64 = None
    if any(s.peso is not None for s in seguimientos):
        plt.figure(figsize=(10, 5))
        pesos = [s.peso for s in seguimientos]
        plt.plot(fechas, pesos, marker='o', color='#8B0000', linewidth=2, label='Peso')
        if peso_ideal:
            plt.axhline(y=peso_ideal, color='green', linestyle='--', linewidth=2, label=f'Ideal: {peso_ideal} kg')
        plt.title("Evolución de Peso", fontsize=14, pad=20)
        plt.xlabel("Fecha", fontsize=12)
        plt.ylabel("Peso (kg)", fontsize=12)
        plt.grid(True, linestyle='--', alpha=0.7)
        plt.legend()
        plt.xticks(rotation=45)
        plt.tight_layout()
        buffer_peso = io.BytesIO()
        plt.savefig(buffer_peso, format='png', dpi=150, bbox_inches='tight')
        plt.close()
        grafica_peso_base64 = base64.b64encode(buffer_peso.getvalue()).decode('utf-8')

    # --- Gráfica IMC ---
    grafica_imc_base64 = None
    if any(s.imc is not None for s in seguimientos):
        plt.figure(figsize=(10, 5))
        imcs = [s.imc for s in seguimientos]
        plt.plot(fechas, imcs, marker='s', color='#6D0000', linewidth=2, label='IMC')
        plt.axhline(y=imc_ideal, color='green', linestyle='--', linewidth=2, label=f'Ideal: {imc_ideal}')
        plt.title("Evolución del IMC", fontsize=14, pad=20)
        plt.xlabel("Fecha", fontsize=12)
        plt.ylabel("IMC", fontsize=12)
        plt.grid(True, linestyle='--', alpha=0.7)
        plt.legend()
        plt.xticks(rotation=45)
        plt.tight_layout()
        buffer_imc = io.BytesIO()
        plt.savefig(buffer_imc, format='png', dpi=150, bbox_inches='tight')
        plt.close()
        grafica_imc_base64 = base64.b64encode(buffer_imc.getvalue()).decode('utf-8')

    # --- Gráfica Circunferencia Abdominal ---
    grafica_circ_abdominal_base64 = None
    if any(s.circunferencia_abdominal is not None for s in seguimientos):
        plt.figure(figsize=(10, 5))
        circ_abdominal = [s.circunferencia_abdominal for s in seguimientos]
        plt.plot(fechas, circ_abdominal, marker='D', color='#A52A2A', linewidth=2, label='Circunferencia')
        plt.axhline(y=circ_ideal, color='green', linestyle='--', linewidth=2, label=f'Ideal: {circ_ideal} cm')
        plt.title("Evolución de Circunferencia Abdominal", fontsize=14, pad=20)
        plt.xlabel("Fecha", fontsize=12)
        plt.ylabel("Circunferencia (cm)", fontsize=12)
        plt.grid(True, linestyle='--', alpha=0.7)
        plt.legend()
        plt.xticks(rotation=45)
        plt.tight_layout()
        buffer_circ = io.BytesIO()
        plt.savefig(buffer_circ, format='png', dpi=150, bbox_inches='tight')
        plt.close()
        grafica_circ_abdominal_base64 = base64.b64encode(buffer_circ.getvalue()).decode('utf-8')

    # --- Render HTML ---
    html_string = render_to_string(
        'reporte_trabajador_pdf.html',
        {
            'trabajador': trabajador,
            'seguimientos': seguimientos,
            'logo_base64': logo_base64,
            'fondo_base64': fondo_base64,
            'grafica_peso_base64': grafica_peso_base64,
            'grafica_imc_base64': grafica_imc_base64,
            'grafica_circ_abdominal_base64': grafica_circ_abdominal_base64,
            'pdf_mode': True
        }
    )

    css_string = f'''
        @page {{
            size: A4;
            margin: 0;
            background-image: url("data:image/png;base64,{fondo_base64}");
            background-size: cover;
            background-position: center center;
            background-repeat: no-repeat;
        }}

        body {{
            font-family: Arial, sans-serif;
            margin: 0;
            padding: 0;
        }}

        .contenedor-principal {{
            background-color: rgba(255, 255, 255, 0.93);
            padding: 2cm;
        }}
    '''

    css = CSS(string=css_string)
    font_config = FontConfiguration()
    html = HTML(string=html_string, base_url=request.build_absolute_uri())
    pdf = html.write_pdf(stylesheets=[css], font_config=font_config)

    response = HttpResponse(pdf, content_type='application/pdf')
    filename = f"reporte_{trabajador.nombre.replace(' ', '_')}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response

def exportar_datos_excel(request):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from datetime import date
    from django.http import HttpResponse
    import io

    from .utils import (
        obtener_desviacion_oms_peso_talla,
        obtener_desviacion_oms_peso_edad,
        obtener_desviacion_oms_talla_edad
    )

    wb = Workbook()

    # ===== ESTILOS =====
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    data_font = Font(size=11)
    data_alignment = Alignment(horizontal="left", vertical="center")
    data_fill_even = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    data_fill_odd = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")

    # ===== HOJA DE PACIENTES =====
    ws_pacientes = wb.active
    ws_pacientes.title = "Pacientes"

    pacientes_qs = Paciente.objects.filter(is_deleted=False).values(
        'nombre', 'curp', 'cai', 'fecha_nacimiento', 'grado', 'grupo', 'peso', 'talla', 'sexo'
    )

    headers = ["Nombre", "CURP", "CAI", "Edad (Años)", "Grado", "Grupo", "Peso (kg)", "Talla (cm)",
               "Peso para Edad", "Talla para Edad", "Peso para Talla"]
    ws_pacientes.append(headers)

    for col in range(1, len(headers) + 1):
        cell = ws_pacientes.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    row_num = 2
    for p in pacientes_qs:
        fecha_nac = p.get('fecha_nacimiento')
        hoy = date.today()
        edad = ''
        if fecha_nac:
            edad = (hoy - fecha_nac).days / 365.25  # años decimales

        peso = p.get('peso') or 0
        talla = p.get('talla') or 0
        sexo = (p.get('sexo') or 'F').upper()

        peso_edad = talla_edad = peso_talla = "ND"

        if peso > 0 and talla > 0 and edad > 0 and sexo in ['M', 'F']:
            try:
                peso_edad = obtener_desviacion_oms_peso_edad(sexo, edad, peso) or "ND"
                talla_edad = obtener_desviacion_oms_talla_edad(sexo, edad, talla) or "ND"
                peso_talla = obtener_desviacion_oms_peso_talla(sexo, talla, peso) or "ND"
            except:
                peso_edad = talla_edad = peso_talla = "Error"

        row = [
            p.get('nombre', ''),
            p.get('curp', ''),
            p.get('cai', ''),
            round(edad, 1) if edad else '',
            p.get('grado', ''),
            p.get('grupo', ''),
            peso,
            talla,
            peso_edad,
            talla_edad,
            peso_talla
        ]
        ws_pacientes.append(row)

        for col in range(1, len(row) + 1):
            cell = ws_pacientes.cell(row=row_num, column=col)
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border
            cell.fill = data_fill_even if row_num % 2 == 0 else data_fill_odd
        row_num += 1

    for col in ws_pacientes.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws_pacientes.column_dimensions[column].width = adjusted_width

    ws_pacientes.freeze_panes = "A2"

    # ===== HOJA DE TRABAJADORES =====
    ws_trabajadores = wb.create_sheet(title="Trabajadores")

    trabajadores_qs = Trabajador.objects.filter(is_deleted=False).values(
        'nombre', 'curp', 'cai', 'cargo', 'departamento', 'imc'
    )

    headers = ["Nombre", "CURP", "CAI", "Puesto", "Departamento", "IMC"]
    ws_trabajadores.append(headers)

    for col in range(1, len(headers) + 1):
        cell = ws_trabajadores.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    row_num = 2
    for t in trabajadores_qs:
        row = [
            t.get('nombre', ''),
            t.get('curp', ''),
            t.get('cai', ''),
            t.get('cargo', ''),
            t.get('departamento', ''),
            round(t.get('imc', 2)) if t.get('imc') else ''
        ]
        ws_trabajadores.append(row)

        for col in range(1, len(row) + 1):
            cell = ws_trabajadores.cell(row=row_num, column=col)
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border
            cell.fill = data_fill_even if row_num % 2 == 0 else data_fill_odd
        row_num += 1

    for col in ws_trabajadores.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws_trabajadores.column_dimensions[column].width = adjusted_width

    ws_trabajadores.freeze_panes = "A2"

    # ===== RESPUESTA =====
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename="reporte_general.xlsx"'

    return response